"""
M&A Monitor — European Public Offers Tracker Generator
=======================================================
Reads the bid premia xlsx export and regenerates index.html with fresh data.

Usage:
    1. Export your bid premia table from M&A Monitor DataBase as xlsx
    2. Place it in this folder (or update XLSX_PATH below)
    3. Run: python update_tracker.py
    4. Commit & push to GitHub: git add . && git commit -m "Data update" && git push

The script will regenerate index.html with the latest deal data.
"""

import openpyxl
import json
import os
import re
from datetime import datetime

# ── CONFIG ──────────────────────────────────────────────────────────────
XLSX_PATH = "bid_premia_latest.xlsx"  # Update this to your export filename
OUTPUT_PATH = "index.html"
TEMPLATE_PATH = "template.html"

# Western European countries only
WESTERN_EUROPE = [
    'United Kingdom', 'France', 'Germany', 'Italy', 'Spain', 'Netherlands',
    'Sweden', 'Norway', 'Denmark', 'Finland', 'Ireland (Republic of)',
    'Luxembourg', 'Portugal', 'Iceland', 'Switzerland', 'Belgium', 'Austria'
]

CODE_MAP = {
    'United Kingdom': 'UK', 'France': 'FR', 'Germany': 'DE', 'Italy': 'IT',
    'Spain': 'ES', 'Netherlands': 'NL', 'Sweden': 'SE', 'Norway': 'NO',
    'Denmark': 'DK', 'Finland': 'FI', 'Ireland (Republic of)': 'IE',
    'Luxembourg': 'LU', 'Portugal': 'PT', 'Iceland': 'IS',
    'Switzerland': 'CH', 'Belgium': 'BE', 'Austria': 'AT'
}


def classify_deal_type(raw_type):
    """Simplify the deal type string into a clean category."""
    raw = str(raw_type or '').lower()
    if 'scheme' in raw:
        return 'Scheme'
    elif 'mand' in raw:
        return 'Mandatory'
    elif 'mbo' in raw or 'mbi' in raw:
        return 'MBO/MBI'
    elif 'merger' in raw:
        return 'Merger'
    elif 'divestiture' in raw:
        return 'Divestiture'
    elif 'partial' in raw:
        return 'Partial'
    else:
        return 'Public Offer'


def clean_adviser_name(raw):
    """
    Clean up adviser names from the database export.
    The raw field often has multiple advisers concatenated with spaces.
    We try to keep them readable and separate with ' / '.
    
    UPDATE THIS FUNCTION as you learn the patterns in your data.
    Add known multi-word firm names to the KNOWN_FIRMS list below.
    """
    if not raw or not raw.strip():
        return ""
    
    # For now, return as-is but trimmed — the adviser names in the DB
    # are hard to split programmatically. You can manually clean them
    # in the OVERRIDES dict below, or improve the splitting logic.
    return raw.strip()


# ── MANUAL OVERRIDES ────────────────────────────────────────────────────
# Use this dict to manually fix target names, bidder names, or adviser 
# formatting for specific deals. Key is the Deal Number from column 1.
# Only include fields you want to override.
#
# Example:
#   "0043861": {"target": "Allfunds Group plc", "tfa": "Citi / Goldman Sachs"}
#
OVERRIDES = {
    # Add your manual name/adviser cleanups here
}


def read_deals(xlsx_path):
    """Read the bid premia xlsx and return a list of deal dicts."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    
    # Find header row (look for 'Deal Number' in column 1)
    header_row = None
    for row in range(1, 15):
        if ws.cell(row, 1).value == 'Deal Number':
            header_row = row
            break
    
    if not header_row:
        raise ValueError("Could not find header row with 'Deal Number' in column 1")
    
    deals = []
    for row in range(header_row + 1, ws.max_row + 1):
        country = ws.cell(row, 6).value
        if country not in WESTERN_EUROPE:
            continue
        
        target = ws.cell(row, 5).value
        if not target:
            continue
        
        deal_num = str(ws.cell(row, 1).value or '')
        premium = ws.cell(row, 20).value  # 1 day before announcement
        ev = ws.cell(row, 15).value or 0
        date_ann = ws.cell(row, 2).value
        date_str = date_ann.strftime('%Y-%m-%d') if date_ann else ''
        
        deal = {
            'target': target.strip(),
            'country': country,
            'code': CODE_MAP.get(country, ''),
            'bidder': (ws.cell(row, 9).value or '').strip(),
            'value': round(ev, 1),
            'premium': round(premium, 1) if premium is not None else None,
            'type': classify_deal_type(ws.cell(row, 11).value),
            'attitude': ws.cell(row, 13).value or 'Friendly',
            'dateAnnounced': date_str,
            'targetFA': clean_adviser_name(ws.cell(row, 30).value),
            'bidderFA': clean_adviser_name(ws.cell(row, 28).value),
        }
        
        # Apply manual overrides if present
        if deal_num in OVERRIDES:
            for key, val in OVERRIDES[deal_num].items():
                if key == 'target': deal['target'] = val
                elif key == 'bidder': deal['bidder'] = val
                elif key == 'tfa': deal['targetFA'] = val
                elif key == 'bfa': deal['bidderFA'] = val
        
        deals.append(deal)
    
    # Sort by equity value descending
    deals.sort(key=lambda x: x['value'], reverse=True)
    return deals


def build_js_array(deals):
    """Convert deals list to a JavaScript array string for embedding."""
    js_deals = []
    for d in deals:
        # Escape any quotes in strings
        def esc(s):
            return s.replace('\\', '\\\\').replace('"', '\\"') if s else ''
        
        premium_str = str(d['premium']) if d['premium'] is not None else 'null'
        
        js_deals.append(
            f'{{t:"{esc(d["target"])}",co:"{esc(d["country"])}",c:"{d["code"]}",'
            f'b:"{esc(d["bidder"])}",v:{d["value"]},p:{premium_str},'
            f'tp:"{d["type"]}",at:"{d["attitude"]}",d:"{d["dateAnnounced"]}",'
            f'tfa:"{esc(d["targetFA"])}",bfa:"{esc(d["bidderFA"])}"}}' 
        )
    
    return 'const deals=[\n' + ',\n'.join(js_deals) + '\n];'


def update_html(deals, template_path, output_path):
    """Read the template HTML, replace the deals array, and write output."""
    with open(template_path, 'r', encoding='utf-8') as f:
        html = f.read()
    
    # Replace the deals array — find the pattern const deals=[...];
    new_js = build_js_array(deals)
    
    # Match from "const deals=[" to the closing "];"
    pattern = r'const deals=\[.*?\];'
    html = re.sub(pattern, new_js, html, flags=re.DOTALL)
    
    # Update the timestamp in the footer
    now = datetime.now().strftime('%#d %b %Y, %H:%M')
    html = re.sub(
        r'Data as at .*? · Source',
        f'Data as at {now} GMT · Source',
        html
    )
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)
    
    print(f"✅ Updated {output_path}")
    print(f"   {len(deals)} Western European deals")
    print(f"   Total value: €{sum(d['value'] for d in deals)/1000:.1f}bn")
    
    # Show deals >€1bn
    big = [d for d in deals if d['value'] >= 1000]
    print(f"   Deals >€1bn: {len(big)}")
    
    hostile = [d for d in deals if d['attitude'] == 'Hostile']
    if hostile:
        print(f"   Hostile: {', '.join(d['target'] for d in hostile)}")


def main():
    # Check xlsx exists
    if not os.path.exists(XLSX_PATH):
        print(f"❌ Cannot find {XLSX_PATH}")
        print(f"   Place your bid premia export in this folder and update XLSX_PATH in the script.")
        print(f"   Current directory: {os.getcwd()}")
        return
    
    # Check template exists — if not, use index.html as template
    template = TEMPLATE_PATH if os.path.exists(TEMPLATE_PATH) else OUTPUT_PATH
    if not os.path.exists(template):
        print(f"❌ Cannot find {template}")
        print(f"   Make sure index.html or template.html exists in this folder.")
        return
    
    print(f"📊 Reading {XLSX_PATH}...")
    deals = read_deals(XLSX_PATH)
    
    print(f"🔨 Generating {OUTPUT_PATH} from {template}...")
    update_html(deals, template, OUTPUT_PATH)
    
    print(f"\n📤 To publish:")
    print(f"   git add .")
    print(f'   git commit -m "Data update {datetime.now().strftime("%d %b %Y")}"')
    print(f"   git push")


if __name__ == '__main__':
    main()
